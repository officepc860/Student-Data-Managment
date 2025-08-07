from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash
import json, os
import pandas as pd
from fpdf import FPDF
from werkzeug.utils import secure_filename
from flask import Flask, session, redirect, url_for, request
import json
from flask import Flask, render_template, request, redirect, url_for, session, flash
from flask import Flask, render_template, request, redirect, session, url_for
from werkzeug.security import generate_password_hash, check_password_hash
import json, os
from flask import Flask, render_template, request, redirect, session, url_for
from werkzeug.security import check_password_hash
import json
import os
from flask import Flask, render_template, request, redirect
from werkzeug.security import generate_password_hash
import json
import os
app = Flask(__name__)
app.secret_key = 'your_secret_key_here'
USERS_FILE = 'users.json'

def load_json(filename):
    if os.path.exists(filename):
        with open(filename, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}
app = Flask(__name__)
app.secret_key = 'your_secret_key_here'
USERS_FILE = 'users.json'

# -------- JSON Utils --------
def load_json(filename):
    if os.path.exists(filename):
        with open(filename, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_json(filename, data):
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4)
app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

# JSON ফাইল থেকে data লোড করার ফাংশন
def load_json(filename):
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

# JSON ফাইলে data সেভ করার ফাংশন
def save_json(filename, data):
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4)

# STUDENTS JSON ফাইলের পাথ
STUDENTS_FILE = 'students.json'
PAYMENT_FILE = 'payment.json'
NEW_PAYMENTS_FILE = 'new_payments.json'  # আপনার ইচ্ছেমতো ফাইল নাম দিতে পারেন
USERS_FILE = 'users.json'


app = Flask(__name__)
app.secret_key = 'test_secret_key_123'

@app.route('/set_session')
def set_session():
    session['user'] = 'testuser'
    return "Session set!"

@app.route('/get_session')
def get_session():
    user = session.get('user')
    return f"Session user is: {user}"
app = Flask(__name__)
app.secret_key = '364770'

STUDENTS_FILE = 'students.json'
PAYMENTS_FILE = 'payment.json'
USERS_FILE = 'users.json'

def get_desktop_folder():
    return os.path.join(os.path.expanduser("~"), "Desktop")

def load_json(file):
    if not os.path.exists(file):
        with open(file, 'w') as f:
            json.dump([] if file == STUDENTS_FILE else {}, f)
    with open(file, 'r') as f:
        return json.load(f)

def save_json(file, data):
    with open(file, 'w') as f:
        json.dump(data, f, indent=4)
from flask import Flask, render_template, request, redirect, session, url_for, jsonify
import sqlite3, hashlib, re

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'


# ---------- INDEX ----------
@app.route('/dashboard')
def dashboard():
    print("Session content:", dict(session))

    if 'user' not in session:
        return redirect(url_for('login'))

    user = session['user']
    user_id = user.get('id')

    students = load_json(STUDENTS_FILE)  # students লোড করো
    user_students = [s for s in students if s.get('user_id') == user_id]  # ফিল্টার করো

    selected_batch = request.args.get('batch', '')
    all_students = user_students[:]

    if selected_batch:
        user_students = [s for s in user_students if s.get('batch') == selected_batch]

    batches = sorted(set(s.get('batch') for s in all_students if s.get('batch')))

    return render_template(
        'dashboard.html',
        students=user_students,
        user=user,
        batches=batches,
        selected_batch=selected_batch
    )


@app.route('/search_student')
def search_student():
    query = request.args.get('query', '').lower()
    batch = request.args.get('batch', '')
    students = load_json(STUDENTS_FILE)
    filtered = [
        s for s in students
        if (query in s['name'].lower() or query in s['roll']) and (s['batch'] == batch if batch else True)
    ]
    batches = sorted(set(s.get('batch') for s in students if s.get('batch')))
    return render_template('dashboard.html', students=filtered, user=session['user'], batches=batches, selected_batch=batch)

@app.route('/add_student', methods=['POST'])
def add_student():
    students = load_json(STUDENTS_FILE)
    roll = request.form['roll']
    batch = request.form['batch']
    name = request.form['name']
    college = request.form['college']
    student_number = request.form.get('student_number', '')
    guardian_number = request.form.get('guardian_number', '')
    
    user_id = session['user']['id']  # এখানে user_id ডিফাইন করো

    # একই user এর মধ্যে roll duplicate আছে কিনা চেক
    if any(s['roll'] == roll and s.get('user_id') == user_id for s in students):
        flash("⚠️ This roll already exists for your account!", "danger")
        return redirect(url_for('dashboard'))

    students.append({
        'roll': roll,
        'batch': batch,
        'name': name,
        'college': college,
        'student_number': student_number,
        'guardian_number': guardian_number,
        'user_id': user_id   # এখানে session user_id যোগ করো
    })
    save_json(STUDENTS_FILE, students)
    flash("✅ Student added successfully!", "success")
    return redirect(url_for('dashboard'))


@app.route('/edit_student/<roll>', methods=['GET', 'POST'])
def edit_student(roll):
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    students = load_json(STUDENTS_FILE)

    # ✅ শুধু নিজের student গুলো দেখার জন্য ফিল্টার
    student = next((s for s in students if s['roll'] == roll and s.get('user_id') == user_id), None)

    if not student:
        flash("⚠️ You are not allowed to edit this student!", "danger")
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        student['batch'] = request.form['batch']
        student['name'] = request.form['name']
        student['college'] = request.form['college']
        student['student_number'] = request.form.get('student_number', '')
        student['guardian_number'] = request.form.get('guardian_number', '')

        save_json(STUDENTS_FILE, students)
        flash("✅ Student updated!", "success")
        return redirect(url_for('dashboard'))

    return render_template('edit_student.html', student=student, user=session['user'])


@app.route('/delete_student/<roll>')
def delete_student(roll):
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    students = load_json(STUDENTS_FILE)

    # শুধু ওই ইউজারের ওই রোল বাদে বাকি সব স্টুডেন্ট রাখবে
    students = [s for s in students if not (s['roll'] == roll and s.get('user_id') == user_id)]

    save_json(STUDENTS_FILE, students)
    flash("🗑️ Student deleted", "info")
    return redirect(url_for('dashboard'))


@app.route('/get_next_roll')
def get_next_roll():
    if 'user' not in session:
        return jsonify({'next_roll': ''})

    user_id = session['user']['id']
    batch = request.args.get('batch', '')
    students = load_json(STUDENTS_FILE)
    rolls = [
        int(s['roll']) for s in students
        if s.get('user_id') == user_id and s.get('batch') == batch and str(s.get('roll', '')).isdigit()
    ]

    next_roll = str(max(rolls) + 1) if rolls else ''
    return jsonify({'next_roll': next_roll})


@app.route('/get_batch_by_roll_prefix')
def get_batch_by_roll_prefix():
    if 'user' not in session:
        return jsonify({'batch': ''})

    user_id = session['user']['id']
    prefix = request.args.get('prefix', '')

    students = load_json(STUDENTS_FILE)
    for s in students:
        if s.get('user_id') == user_id and str(s.get('roll', '')).startswith(prefix):
            return jsonify({'batch': s.get('batch', '')})

    return jsonify({'batch': ''})


from flask import jsonify

@app.route('/payment', methods=['GET', 'POST'])
def payment():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    all_students = load_json(STUDENTS_FILE)
    all_payments = load_json(PAYMENTS_FILE)
    all_new_payments = load_json(NEW_PAYMENTS_FILE)

    user_students = [s for s in all_students if s.get('user_id') == user_id]
    students_dict = {s['roll']: s for s in user_students}
    user_payments = all_payments.get(user_id, {})

    user_new_payments_raw = all_new_payments.get(user_id)
    if user_new_payments_raw is None:
        user_new_payments_list = []
    elif isinstance(user_new_payments_raw, list):
        user_new_payments_list = user_new_payments_raw
    elif isinstance(user_new_payments_raw, dict):
        user_new_payments_list = list(user_new_payments_raw.values())
    else:
        user_new_payments_list = []

    batches = sorted(set(s.get('batch') for s in user_students if s.get('batch')))
    selected_batch = request.args.get('batch', '')
    selected_status = request.args.get('status', '')

    roll = request.args.get('roll') or request.form.get('roll')
    student = students_dict.get(roll)

    payment_info = {
        "date": "",
        "memo_no": "",
        "receipt_no": "",
        "course": "",
        "total_payment": "",
        "previous_payment": 0,
        "new_payment": 0,
        "discount": 0,
        "due": 0,
        "reference": "",
        "status": "Due"
    }

    if roll and student:
        if request.method == 'POST':
            date = request.form.get('date', '').strip()
            memo_no = request.form.get('memo_no', '').strip()
            session['last_memo_no'] = memo_no

            if len(memo_no) >= 2:
                memo_prefix = memo_no[:2]
                for entry in user_new_payments_list:
                    existing_memo = entry.get('memo_no', '')
                    if existing_memo and existing_memo[:2] != memo_prefix and not existing_memo.startswith(memo_prefix):
                        continue
                    if existing_memo and existing_memo[:2] != memo_prefix and existing_memo != memo_no:
                        flash("❌ This Memo No prefix conflicts with another memo. Please choose unique starting digits.", "danger")
                        return redirect(url_for('payment', roll=roll, batch=selected_batch, status=selected_status))

            receipt_no_input = request.form.get('receipt_no', '').strip()
            total = int(request.form['total_payment'])
            previous = int(request.form.get('previous_payment', 0))
            new_payment = int(request.form.get('new_payment', 0))
            discount = int(request.form.get('discount', 0))
            course = request.form.get('course', '').strip()
            reference = request.form.get('reference', '').strip()

            memo_payments = [entry for entry in user_new_payments_list if entry.get('memo_no') == memo_no]
            if memo_payments:
                try:
                    max_receipt_no = max(int(p.get('receipt_no', 0)) for p in memo_payments)
                except ValueError:
                    max_receipt_no = 0
                receipt_no = str(max_receipt_no + 1)
            else:
                receipt_no = receipt_no_input

            for entry in user_new_payments_list:
                if entry.get('receipt_no') == receipt_no and entry.get('memo_no') == memo_no:
                    flash(f"❌ Receipt No '{receipt_no}' already exists under this Memo No. Please use a unique receipt number.", "danger")
                    return redirect(url_for('payment', roll=roll, batch=selected_batch, status=selected_status))

            updated_previous = previous + new_payment
            due = total - (updated_previous + discount)
            status = "Paid" if due <= 0 else "Due"

            user_payments[roll] = {
                "name": student['name'],
                "batch": student['batch'],
                "date": date,
                "course": course,
                "total_payment": total,
                "previous_payment": updated_previous,
                "discount": discount,
                "due": due,
                "reference": reference,
                "status": status
            }
            all_payments[user_id] = user_payments
            save_json(PAYMENTS_FILE, all_payments)

            user_new_payments_list.append({
                "roll": roll,
                "new_payment": new_payment,
                "date": date,
                "memo_no": memo_no,
                "receipt_no": receipt_no,
                "course": course
            })
            all_new_payments[user_id] = user_new_payments_list
            save_json(NEW_PAYMENTS_FILE, all_new_payments)

            flash("✅ Payment updated", "success")
            return redirect(url_for('payment', roll=roll, batch=selected_batch, status=selected_status))

        else:
            info = user_payments.get(roll, {})
            memo_no_get = request.args.get('memo_no')
            memo_no = memo_no_get if memo_no_get is not None else session.get('last_memo_no', '')

            memo_payments = [entry for entry in user_new_payments_list if entry.get('memo_no') == memo_no]
            if memo_payments:
                try:
                    max_receipt_no = max(int(p.get('receipt_no', 0)) for p in memo_payments)
                    next_receipt_no = str(max_receipt_no + 1)
                except ValueError:
                    next_receipt_no = ''
            else:
                next_receipt_no = info.get('receipt_no', '')

            payment_info = {
                "date": info.get("date", ""),
                "memo_no": memo_no,
                "receipt_no": next_receipt_no,
                "course": info.get("course", ""),
                "total_payment": info.get("total_payment", 0),
                "previous_payment": info.get("previous_payment", 0),
                "new_payment": 0,
                "due": info.get("due", 0),
                "discount": info.get("discount", 0),
                "reference": info.get("reference", ""),
                "status": "Paid" if info.get("due", 0) <= 0 else "Due"
            }

    filtered_students = [s for s in user_students if not selected_batch or s.get('batch') == selected_batch]
    filtered_payments = {
        roll: p for roll, p in user_payments.items()
        if (not selected_batch or p.get('batch') == selected_batch) and
           (not selected_status or p.get('status') == selected_status)
    }

    return render_template(
        'payment.html',
        students=filtered_students,
        payments=user_payments,
        batches=batches,
        selected_batch=selected_batch,
        selected_status=selected_status,
        student=student,
        payment=payment_info,
        roll=roll,
        filtered_payments=filtered_payments,
        user=session['user']
    )



@app.route('/download_memo_excel')
def download_memo_excel():
    if 'user' not in session:
        flash("Please login first", "warning")
        return redirect(url_for('login'))

    user_id = session['user']['id']
    memo_no = request.args.get('memo_no', '').strip()
    if not memo_no:
        flash("Memo No is required to download Excel.", "danger")
        return redirect(url_for('payment'))

    all_new_payments = load_json(NEW_PAYMENTS_FILE)
    all_students = load_json(STUDENTS_FILE)
    all_payments = load_json(PAYMENTS_FILE)

    user_new_payments_list = all_new_payments.get(user_id, [])

    # Filter only those payments with matching memo_no
    filtered_payments = [p for p in user_new_payments_list if p.get('memo_no', '') == memo_no]

    if not filtered_payments:
        flash(f"No payments found for Memo No: {memo_no}", "info")
        return redirect(url_for('payment'))

    # Prepare data rows for Excel
    data = []
    # For student info mapping by roll
    student_dict = {s['roll']: s for s in all_students if s.get('user_id') == user_id}

    for pay in filtered_payments:
        roll = pay.get('roll')
        student = student_dict.get(roll, {})
        # Original payment info (previous, total, discount, due) থেকে নিতে পারো
        payment_info = all_payments.get(user_id, {}).get(roll, {})

        data.append({
            "Roll": roll,
            "Name": student.get('name', ''),
            "Batch": student.get('batch', ''),
            "Date": pay.get('date', ''),
            "Memo No": pay.get('memo_no', ''),
            "Receipt No": pay.get('receipt_no', ''),
            "Course": pay.get('course', ''),
            "New Payment": pay.get('new_payment', 0),
            "Total Payment": payment_info.get('total_payment', 0),
            "Previous Payment": payment_info.get('previous_payment', 0),
            "Discount": payment_info.get('discount', 0),
            "Due": payment_info.get('due', 0),
            "Status": payment_info.get('status', 'Due')
        })

    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=f'Memo_{memo_no}')

        workbook = writer.book
        worksheet = writer.sheets[f'Memo_{memo_no}']

        # Adjust column widths
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            col_letter = column_cells[0].column_letter
            worksheet.column_dimensions[col_letter].width = length + 2

        # Header formatting
        from openpyxl.styles import PatternFill, Font
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

    output.seek(0)

    filename = f"memo_{memo_no}_payments.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
import io
import pandas as pd
from flask import send_file, flash, redirect, url_for, session
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

import io
import pandas as pd
from flask import send_file, flash, redirect, url_for, session
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

@app.route('/download_memo_list_excel')
def download_memo_list_excel():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    memo_filter = request.args.get('memo_no', '').strip()

    all_new_payments = load_json(NEW_PAYMENTS_FILE)
    user_new_payments_list = all_new_payments.get(user_id, [])

    if memo_filter:
        user_new_payments_list = [
            p for p in user_new_payments_list if str(p.get('memo_no')) == memo_filter
        ]

    students = load_students_dict()

    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Memo Receipts"

    # Header row — same as PDF
    headers = ["Date", "Memo No", "Receipt No", "Roll", "Name", "Course", "Amount"]
    ws.append(headers)

    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    total_amount = 0

    for entry in user_new_payments_list:
        roll = entry.get("roll", "")
        name = students.get(roll, {}).get("name", "")
        amount = float(entry.get("new_payment", 0))

        row = [
            entry.get("date", entry.get("payment_date", "")),
            entry.get("memo_no", ""),
            entry.get("receipt_no", ""),
            roll,
            name,
            entry.get("course", ""),
            amount
        ]
        ws.append(row)
        total_amount += amount

    # Add total row at the bottom
    total_row_index = ws.max_row + 1
    ws[f"F{total_row_index}"] = "Total:"
    ws[f"F{total_row_index}"].font = Font(bold=True)
    ws[f"G{total_row_index}"] = total_amount
    ws[f"G{total_row_index}"].font = Font(bold=True)

    # Optional: Adjust column width
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Save to BytesIO and return
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"memo_receipt_list_{memo_filter or 'all'}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')




@app.route('/get_next_receipt_no')
def get_next_receipt_no():
    if 'user' not in session:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401

    memo_no = request.args.get('memo_no', '').strip()
    roll = request.args.get('roll', '').strip()
    if not memo_no or not roll:
        return jsonify({"status": "error", "message": "Missing parameters"}), 400

    user_id = session['user']['id']
    all_new_payments = load_json(NEW_PAYMENTS_FILE)
    user_new_payments_list = all_new_payments.get(user_id, [])

    # Memo prefix conflict check (first 2 digits)
    if len(memo_no) >= 2:
        memo_prefix = memo_no[:2]
        for entry in user_new_payments_list:
            existing_memo = entry.get('memo_no', '')
            if existing_memo and existing_memo[:2] == memo_prefix and existing_memo != memo_no:
                return jsonify({
                    "status": "error",
                    "message": "This Memo No prefix conflicts with another memo. Please choose unique starting digits."
                })

    memo_payments = [entry for entry in user_new_payments_list if entry.get('memo_no') == memo_no]

    if memo_payments:
        try:
            max_receipt_no = max(int(p.get('receipt_no', 0)) for p in memo_payments)
        except ValueError:
            max_receipt_no = 0
        next_receipt_no = str(max_receipt_no + 1)
    else:
        next_receipt_no = "1"

    return jsonify({"status": "success", "receipt_no": next_receipt_no})


import json
import os

def load_json(filepath):
    if not os.path.exists(filepath):
        return {}
    with open(filepath, 'r', encoding='utf-8') as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return {}

def save_json(filepath, data):
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4)

from math import ceil
from flask import request

@app.route('/memo_list')
def memo_list():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    all_new_payments = load_json(NEW_PAYMENTS_FILE)
    all_students = load_json(STUDENTS_FILE)
    user_new_payments_list = all_new_payments.get(user_id, [])
    user_students = {s['roll']: s for s in all_students if s.get('user_id') == user_id}

    search_memo_no = request.args.get('memo_no', '').strip()
    search_roll = request.args.get('roll', '').strip()
    search_name = request.args.get('name', '').strip()
    page = int(request.args.get('page', 1))
    per_page = 50  # প্রতি পেজে ৫০টি

    # Filter memos
    filtered_memos = []
    for entry in user_new_payments_list:
        if search_memo_no and search_memo_no not in entry.get('memo_no', ''):
            continue
        if search_roll and search_roll != entry.get('roll', ''):
            continue
        
        student = user_students.get(entry.get('roll'))
        student_name = student['name'] if student else ''

        if search_name and search_name.lower() not in student_name.lower():
            continue

        memo = entry.copy()
        memo['name'] = student_name
        memo['amount'] = entry.get('new_payment', 0)
        filtered_memos.append(memo)

    # Sort by memo_no (string) and then receipt_no (int)
    def memo_sort_key(x):
        memo_no = x.get('memo_no', '')
        receipt_no = int(x.get('receipt_no', 0)) if x.get('receipt_no', '0').isdigit() else 0
        return (memo_no, receipt_no)

    filtered_memos.sort(key=memo_sort_key)

    # Pagination
    total = len(filtered_memos)
    total_pages = ceil(total / per_page)
    start = (page - 1) * per_page
    end = start + per_page
    paginated_memos = filtered_memos[start:end]

    return render_template(
        'memo_list.html',
        memos=paginated_memos,
        page=page,
        total_pages=total_pages,
        search_memo_no=search_memo_no,
        search_roll=search_roll,
        search_name=search_name
    )

from flask import send_file
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
import io
import math

from flask import send_file
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import cm
import io

from flask import send_file, session, redirect, url_for
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import cm
import io

from flask import send_file, session, redirect, url_for
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import cm
import io
def load_students_dict():
    students_list = load_json(STUDENTS_FILE)  # এখানে students.json যেটা লোড করবেন
    # লিস্ট থেকে dict বানানো, যেখানে key = roll, value = student dict
    return {student['roll']: student for student in students_list}

def load_students_dict():
    students_list = load_json(STUDENTS_FILE)  # এখানে students.json যেটা লোড করবেন
    # লিস্ট থেকে dict বানানো, যেখানে key = roll, value = student dict
    return {student['roll']: student for student in students_list}

from flask import send_file, session, redirect, url_for
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import cm
import io


@app.route('/export_memo_receipt_pdf')
def export_memo_receipt_pdf():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    memo_filter = request.args.get('memo_no', '').strip()

    all_new_payments = load_json(NEW_PAYMENTS_FILE)
    user_new_payments_list = all_new_payments.get(user_id, [])

    if memo_filter:
        user_new_payments_list = [
            p for p in user_new_payments_list if str(p.get('memo_no')) == memo_filter
        ]

    students = load_students_dict()

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)
    margin = 1.5 * cm
    row_height = 20

    headers = ["Date", "Memo No", "Receipt No", "Roll", "Name", "Course", "Amount"]
    col_widths = [3*cm, 3*cm, 3*cm, 3*cm, 8*cm, 4.5*cm, 3*cm]
    x_positions = [margin]
    for w in col_widths[:-1]:
        x_positions.append(x_positions[-1] + w)
    end_x = x_positions[-1] + col_widths[-1]

    grand_total = 0
    page_total = 0
    cumulative_total = 0
    y = height - margin
    row_count = 0
    current_memo = None

    def draw_table_header():
        nonlocal y
        y = height - margin
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawCentredString(width / 2, y, "Memo & Receipt Payment Report")
        y -= row_height * 1.5
        pdf.setFont("Helvetica-Bold", 10)
        for i, header in enumerate(headers):
            col_center = x_positions[i] + col_widths[i] / 2
            pdf.drawCentredString(col_center, y + 5, header)
        pdf.line(margin, y + row_height, end_x, y + row_height)
        pdf.line(margin, y, end_x, y)
        for x in x_positions:
            pdf.line(x, y, x, y + row_height)
        pdf.line(end_x, y, end_x, y + row_height)
        y -= row_height

    draw_table_header()

    for entry in user_new_payments_list:
        entry_memo = entry.get("memo_no", "")
        if (
            row_count == 50 or
            y < margin + 3 * row_height or
            (current_memo is not None and current_memo != entry_memo)
        ):
            # Draw page total
            pdf.setFont("Helvetica-Bold", 10)
            pdf.drawRightString(end_x - 2, y + 5, f"Page Total: {page_total:.2f}")
            cumulative_total += page_total
            pdf.drawRightString(end_x - 2, y - 10, f"Grand Total So Far: {cumulative_total:.2f}")
            grand_total = cumulative_total
            page_total = 0
            row_count = 0
            pdf.showPage()
            draw_table_header()

        current_memo = entry_memo
        roll = entry.get("roll", "")
        name = students.get(roll, {}).get("name", "") if roll else ""
        raw_date = entry.get("date", "")

        try:
            formatted_date = datetime.strptime(raw_date, "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            formatted_date = raw_date

        row_data = [
            formatted_date,
            entry.get("memo_no", ""),
            entry.get("receipt_no", ""),
            roll,
            name,
            entry.get("course", ""),
            f"{float(entry.get('new_payment', 0)):.2f}"
        ]

        pdf.setFont("Helvetica", 10)
        for i, value in enumerate(row_data):
            col_center = x_positions[i] + col_widths[i] / 2
            pdf.drawCentredString(col_center, y + 5, str(value))

        pdf.line(margin, y, end_x, y)
        for x in x_positions:
            pdf.line(x, y, x, y + row_height)
        pdf.line(end_x, y, end_x, y + row_height)

        y -= row_height
        amount = float(entry.get("new_payment", 0))
        page_total += amount
        row_count += 1

    # Last page totals
    pdf.line(margin, y, end_x, y)
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawRightString(end_x - 2, y + 5, f"Page Total: {page_total:.2f}")
    cumulative_total += page_total
    pdf.drawRightString(end_x - 2, y - 10, f"Grand Total: {cumulative_total:.2f}")
    pdf.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="memo_receipt_report.pdf",
        mimetype='application/pdf'
    )

@app.route('/memo/edit/<memo_no>/<receipt_no>', methods=['GET', 'POST'])
def edit_memo_receipt(memo_no, receipt_no):
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    all_new_payments = load_json(NEW_PAYMENTS_FILE)
    user_entries = all_new_payments.get(user_id, [])

    # Entry খোঁজা
    entry = next((e for e in user_entries if e.get('memo_no') == memo_no and str(e.get('receipt_no')) == str(receipt_no)), None)

    if not entry:
        flash("❌ Memo entry not found", "danger")
        return redirect(url_for('memo_list'))

    if request.method == 'POST':
        entry['date'] = request.form.get('date', '').strip()
        entry['memo_no'] = request.form.get('memo_no', '').strip()
        entry['receipt_no'] = request.form.get('receipt_no', '').strip()
        entry['course'] = request.form.get('course', '').strip()
        entry['new_payment'] = float(request.form.get('new_payment', 0))

        save_json(NEW_PAYMENTS_FILE, all_new_payments)
        flash("✅ Memo updated successfully", "success")
        return redirect(url_for('memo_list'))

    return render_template('edit_memo.html', entry=entry)

@app.route('/memo/delete/<memo_no>/<receipt_no>', methods=['POST'])
def delete_memo_receipt(memo_no, receipt_no):
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    all_new_payments = load_json(NEW_PAYMENTS_FILE)

    user_entries = all_new_payments.get(user_id, [])
    updated_entries = [
        entry for entry in user_entries
        if not (str(entry.get('memo_no')) == str(memo_no) and str(entry.get('receipt_no')) == str(receipt_no))
    ]

    if len(updated_entries) == len(user_entries):
        flash("❌ Memo receipt not found or already deleted.", "warning")
    else:
        all_new_payments[user_id] = updated_entries
        save_json(NEW_PAYMENTS_FILE, all_new_payments)
        flash("🗑️ Memo receipt deleted successfully.", "success")

    return redirect(url_for('memo_list'))

import os
import pandas as pd
from flask import flash, redirect, url_for, request, send_file
from io import BytesIO

from flask import flash, redirect, url_for, request, send_file
from io import BytesIO
import os
import pandas as pd

@app.route('/export_excel_custom')
def export_excel_custom():
    students = load_json(STUDENTS_FILE)
    batch = request.args.get('batch')
    download = request.args.get('download') == '1'  # download=1 পেলে ডাউনলোড মোড

    if batch:
        students = [s for s in students if s.get('batch') == batch]

    if not students:
        flash("⚠️ No student data to export", "warning")
        return redirect(url_for('dashboard', batch=batch))

    df = pd.DataFrame(students)

    if download:
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Students')
        output.seek(0)
        filename = f"students_{batch or 'all'}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    else:
        folder = get_desktop_folder()
        if not folder:
            flash("❗ Desktop folder not found.", "warning")
            return redirect(url_for('dashboard', batch=batch))

        excel_path = os.path.join(folder, f"students_{batch or 'all'}.xlsx")
        try:
            df.to_excel(excel_path, index=False)
            flash(f"✅ Excel exported to: {excel_path}", "success")
        except Exception as e:
            flash(f"❌ Excel export failed: {str(e)}", "danger")

        return redirect(url_for('dashboard', batch=batch))


@app.route('/view_excel')
def view_excel():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = str(session['user']['id'])
    batch = request.args.get('batch')

    students = load_json(STUDENTS_FILE)
    students = [s for s in students if str(s.get('user_id')) == user_id]

    if batch and batch.lower() != 'all':
        students = [s for s in students if s.get('batch') == batch]

    return render_template('view_excel.html', students=students, batch=batch)


from flask import send_file
from io import BytesIO

from flask import send_file, session, redirect, url_for, request
from fpdf import FPDF
from io import BytesIO

@app.route('/export_pdf_custom')
def export_pdf_custom():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = str(session['user']['id'])
    batch = request.args.get('batch')
    download = request.args.get('download') == '1'

    print("Batch:", batch)
    print("Download:", download)

    students = load_json(STUDENTS_FILE)
    students = [s for s in students if str(s.get('user_id')) == user_id]

    if batch and batch.lower() != 'all':
        students = [s for s in students if s.get('batch') == batch]

    students = load_json(STUDENTS_FILE)
    students = [s for s in students if s.get('user_id') == user_id]
    if batch:
        students = [s for s in students if s.get('batch') == batch]

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 40)  # ডিফল্ট Arial ফন্ট, একটু মোটা (bold)
    pdf.cell(190, 15, 'RETINA', ln=True, align='C')
    pdf.set_font("Arial", size=15)
    pdf.cell(190, 10, txt="Student List", ln=True, align='C')
    pdf.ln(5)

    col_widths = [10, 15, 15, 60, 30, 30, 30]
    row_height = 8
    headers = ["SL", "Roll", "Batch", "Name", "College", "Student No", "Guardian No"]

    def add_table_header():
        pdf.set_font("Arial", 'B', 10)
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], row_height, header, 1, 0, 'C')
        pdf.ln()

    add_table_header()
    pdf.set_font("Arial", '', 10)

    for idx, s in enumerate(students, start=1):
        if pdf.get_y() > 270:
            pdf.add_page()
            add_table_header()
            pdf.set_font("Arial", '', 10)

        data = [
            str(idx),
            s.get("roll", ""),
            s.get("batch", ""),
            s.get("name", ""),
            s.get("college", ""),
            s.get("student_number", ""),
            s.get("guardian_number", "")
        ]

        for i, txt in enumerate(data):
            y_before = pdf.get_y()
            x_before = pdf.get_x()

            pdf.cell(col_widths[i], row_height, '', 1)
            v_offset = (row_height - 5) / 2
            pdf.set_xy(x_before, y_before + v_offset)
            pdf.cell(col_widths[i], 5, txt, 0, 0, 'C')
            pdf.set_xy(x_before + col_widths[i], y_before)

        pdf.ln(row_height)

    pdf_data = pdf.output(dest='S').encode('latin1')
    pdf_output = BytesIO(pdf_data)

    return send_file(
        pdf_output,
        mimetype='application/pdf',
        as_attachment=download,
        download_name=f"students_{batch or 'all'}.pdf"
    )


@app.route('/import_excel', methods=['POST'])
def import_excel():
    if 'user' not in session:
        flash("Please log in first.", "warning")
        return redirect(url_for('login'))

    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        flash("⚠️ Please upload a valid Excel (.xlsx) file!", "danger")
        return redirect(url_for('dashboard'))
    try:
        df = pd.read_excel(file)
        required_columns = {'name', 'roll', 'batch', 'student_number', 'guardian_number'}
        if not required_columns.issubset(df.columns):
            flash("⚠️ Excel file is missing required columns!", "danger")
            return redirect(url_for('dashboard'))

        students = load_json(STUDENTS_FILE)
        existing_rolls = {s['roll'] for s in students}
        added = 0
        user_id = session['user']['id']  # ✅ current user ID

        for _, row in df.iterrows():
            if str(row['roll']) not in existing_rolls:
                students.append({
                    'roll': str(row.get('roll', '')),
                    'batch': row.get('batch', ''),
                    'name': row.get('name', ''),
                    'college': row.get('college', ''),
                    'student_number': str(row.get('student_number', '')),
                    'guardian_number': str(row.get('guardian_number', '')),
                    'user_id': user_id  # ✅ attach user_id
                })
                added += 1

        save_json(STUDENTS_FILE, students)
        flash(f"✅ {added} new student(s) imported successfully!", "success")
    except Exception as e:
        flash(f"❌ Import failed: {str(e)}", "danger")
    return redirect(url_for('dashboard'))


@app.route('/import_payment_excel', methods=['POST'])
def import_payment_excel():
    if 'user' not in session:
        flash("Please log in first.", "warning")
        return redirect(url_for('login'))

    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        flash("⚠️ Please upload a valid Excel (.xlsx) file!", "danger")
        return redirect(url_for('dashboard'))

    try:
        df = pd.read_excel(file)
        required_columns = {'roll', 'total_payment', 'previous_payment', 'discount', 'reference'}
        if not required_columns.issubset(df.columns):
            flash("⚠️ Excel must have columns: roll, total_payment, previous_payment, discount, and reference!", "danger")
            return redirect(url_for('dashboard'))

        students = load_json(STUDENTS_FILE)
        payments = load_json(PAYMENTS_FILE)
        user_id = session['user']['id']

        if user_id not in payments:
            payments[user_id] = {}

        student_dict = {s['roll']: s for s in students if s.get('user_id') == user_id}
        updated = 0

        for _, row in df.iterrows():
            roll = str(row.get('roll', '')).strip()
            if roll in student_dict:
                try:
                    total = int(row.get('total_payment', 0) or 0)
                    previous = int(row.get('previous_payment', 0) or 0)
                    discount = int(row.get('discount', 0) or 0)
                    reference = str(row.get('reference', '')).strip()
                    due = total - (previous + discount)
                    status = "Paid" if due <= 0 else "Due"

                    payments[user_id][roll] = {
                        "name": student_dict[roll]['name'],
                        "batch": student_dict[roll]['batch'],
                        "total_payment": total,
                        "previous_payment": previous,
                        "discount": discount,
                        "due": due,
                        "status": status,
                        "reference": reference,
                    }
                    updated += 1
                except Exception as e:
                    flash(f"⚠️ Error processing roll {roll}: {str(e)}", "warning")

        save_json(PAYMENTS_FILE, payments)
        flash(f"✅ Payment data imported for {updated} student(s)!", "success")

    except Exception as e:
        flash(f"❌ Import failed: {str(e)}", "danger")

    return redirect(url_for('dashboard'))


@app.route('/view_payment_excel')
def view_payment_excel():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    batch = request.args.get('batch', '')
    status = request.args.get('status', '')

    payments = load_json(PAYMENTS_FILE)
    user_payments = payments.get(str(user_id), {})

    filtered = {
        roll: p for roll, p in user_payments.items()
        if (not batch or p.get('batch') == batch) and
           (not status or p.get('status') == status)
    }

    data = []
    for roll, p in filtered.items():
        data.append({
            'roll': roll,
            'name': p.get('name', ''),
            'batch': p.get('batch', ''),
            'total_payment': p.get('total_payment', 0),
            'previous_payment': p.get('previous_payment', 0),
            'discount': p.get('discount', 0),
            'reference': p.get('reference', 0),
            'due': p.get('due', 0),
            'status': p.get('status', '')
        })

    return render_template('view_payment_excel.html', payments=data, batch=batch, status=status)


# ---------- EXPORT PAYMENT EXCEL ----------
@app.route('/export_payment_excel')
def export_payment_excel():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    batch = request.args.get('batch', '')
    status = request.args.get('status', '')
    download = request.args.get('download', '0') == '1'

    payments = load_json(PAYMENTS_FILE)
    user_payments = payments.get(str(user_id), {})

    filtered = {
        roll: p for roll, p in user_payments.items()
        if (not batch or p.get('batch') == batch) and
           (not status or p.get('status') == status)
    }

    if not filtered:
        flash("⚠️ No payment data found for export!", "warning")
        return redirect(url_for('dashboard'))

    data = []
    for roll, p in filtered.items():
        data.append({
            'Roll': roll,
            'Name': p.get('name', ''),
            'Batch': p.get('batch', ''),
            'Total Payment': p.get('total_payment', 0),
            'Previous Payment': p.get('previous_payment', 0),
            'Discount': p.get('discount', 0),
            'Reference': p.get('reference', 0),
            'Due': p.get('due', 0),
            'Status': p.get('status', '')
        })

    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Payments')
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=download,
        download_name=f"payment_report_{batch or 'all'}_{status or 'all'}.xlsx"
    )


# ---------- EXPORT PAYMENT PDF ----------
@app.route('/export_payment_pdf')
def export_payment_pdf():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = str(session['user']['id'])
    batch = request.args.get('batch', '')
    status_filter = request.args.get('status', '')
    download = request.args.get('download', '0') == '1'

    payments = load_json(PAYMENTS_FILE)
    user_payments = payments.get(user_id, {})

    filtered = {
        roll: p for roll, p in user_payments.items()
        if (not batch or p.get('batch') == batch)
        and (not status_filter or p.get('status') == status_filter)
    }

    if not filtered:
        flash("⚠️ No payment data found for this batch and status!", "warning")
        return redirect(url_for('dashboard'))

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font('Arial', 'B', 40)  # ডিফল্ট Arial ফন্ট, একটু মোটা (bold)
    pdf.cell(190, 15, 'RETINA', ln=True, align='C')
    pdf.set_font("Arial", size=15)
    pdf.cell(190, 10, txt=f"Payment Report - Batch: {batch or 'All'} | Status: {status_filter or 'All'}", ln=True, align='C')
    pdf.ln(5)

    headers = ["Roll", "Name", "Batch", "Total", "Prev", "Discount", "Due", "Ref", "Status"]
    col_widths = [15, 50, 15, 20, 20, 20, 15, 20, 15]

    pdf.set_font("Arial", 'B', 10)
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 8, txt=h, border=1, align='C')
    pdf.ln()

    pdf.set_font("Arial", size=10)
    for roll, p in filtered.items():
        row = [
            roll,
            p.get("name", ""),
            p.get("batch", ""),
            str(p.get("total_payment", 0)),
            str(p.get("previous_payment", 0)),
            str(p.get("discount", 0)),
            str(p.get("reference", "")),
            str(p.get("due", 0)),
            p.get("status", "")
        ]
        for i, val in enumerate(row):
            align = 'L' if i == 1 else 'C'
            pdf.cell(col_widths[i], 8, txt=val[:30] if i == 1 else val[:12], border=1, align=align)
        pdf.ln()

    pdf_data = pdf.output(dest='S').encode('latin1')
    pdf_io = BytesIO(pdf_data)

    return send_file(
        pdf_io,
        mimetype='application/pdf',
        as_attachment=download,
        download_name=f"payment_report_{batch or 'all'}_{status_filter or 'all'}.pdf"
    )

@app.route('/delete_payment/<roll>')
def delete_payment(roll):
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    all_payments = load_json(PAYMENTS_FILE)

    # এই পেমেন্ট গুলো user_id key দ্বারা সংরক্ষিত dict হিসেবে ধরে নিচ্ছি
    user_payments = all_payments.get(user_id, {})

    if roll in user_payments:
        del user_payments[roll]
        all_payments[user_id] = user_payments
        save_json(PAYMENTS_FILE, all_payments)
        flash(f"🗑️ Payment deleted for Roll {roll}", "success")

    return redirect(url_for('payment'))



# ---------- AUTH ----------
import uuid  # এইটা অবশ্যই import করতে হবে উপরে

import uuid  # Ensure this is at the top of your file

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email'].strip().lower()
        phone = request.form['phone'].strip()
        raw_password = request.form['password'].strip()

        if not email and not phone:
            return render_template('register.html', error="Email বা Phone যেকোনো একটি দিতে হবে।")

        users = load_json(USERS_FILE)

        # Unique check: email বা phone আগেই আছে কিনা
        for user in users.values():
            if email and user.get('email') == email:
                return render_template('register.html', error="Email আগেই ব্যবহার হয়েছে।")
            if phone and user.get('phone') == phone:
                return render_template('register.html', error="Phone আগেই ব্যবহার হয়েছে।")

        # Password hash করা
        password = generate_password_hash(raw_password)

        # নতুন user তৈরি - UUID দিয়ে unique ID
        user_id = str(uuid.uuid4())
        users[user_id] = {
            "name": name,
            "email": email,
            "phone": phone,
            "password": password
        }

        save_json(USERS_FILE, users)
        return redirect('/login')

    return render_template('register.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        login_id = request.form['login'].strip().lower()
        password = request.form['password'].strip()

        users = load_json(USERS_FILE)
        for user_id, user in users.items():
            if user.get('email') == login_id or user.get('phone') == login_id:
                if check_password_hash(user['password'], password):
                    session['user'] = {
                        'id': user_id,
                        'is_admin': user.get('is_admin', False),
                        'name': user.get('name', 'User')  # Add this line
                    }
                    return redirect(url_for('dashboard'))
                break
        return render_template('login.html', error="Invalid credentials")
    return render_template('login.html')




@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

@app.route('/forgot', methods=['GET', 'POST'])
def forgot():
    if request.method == 'POST':
        login_id = request.form['login'].strip().lower()
        users = load_json(USERS_FILE)
        for user_id, user in users.items():
            if user.get('email') == login_id or user.get('phone') == login_id:
                return redirect(url_for('reset_password', user_id=user_id))
        return render_template('forgot_password.html', error="User not found")
    return render_template('forgot_password.html')

@app.route('/reset/<user_id>', methods=['GET', 'POST'])
def reset_password(user_id):
    users = load_json(USERS_FILE)
    if user_id not in users:
        return redirect('/forgot')

    if request.method == 'POST':
        new_password = request.form['password'].strip()
        users[user_id]['password'] = generate_password_hash(new_password)
        save_json(USERS_FILE, users)
        return redirect('/login')

    return render_template('reset_password.html')

# -------- ADMIN PANEL --------
@app.route('/admin')
def admin():
    if not session.get('user') or not session['user'].get('is_admin'):
        return redirect('/')
    users = load_json(USERS_FILE)
    users_list = [
        {
            'id': uid,
            'name': u.get('name'),
            'email': u.get('email'),
            'phone': u.get('phone'),
            'is_admin': u.get('is_admin', False)
        }
        for uid, u in users.items()
    ]
    return render_template('admin.html', users=users_list)

@app.route('/admin/delete/<user_id>')
def admin_delete(user_id):
    if not session.get('user') or not session['user'].get('is_admin'):
        return redirect('/')
    users = load_json(USERS_FILE)
    if user_id in users:
        users.pop(user_id)
        save_json(USERS_FILE, users)
    return redirect('/admin')

@app.route('/admin/promote/<user_id>')
def admin_promote(user_id):
    if not session.get('user') or not session['user'].get('is_admin'):
        return redirect('/')
    users = load_json(USERS_FILE)
    if user_id in users:
        users[user_id]['is_admin'] = True
        save_json(USERS_FILE, users)
    return redirect('/admin')

@app.route('/admin/demote/<user_id>')
def admin_demote(user_id):
    if not session.get('user') or not session['user'].get('is_admin'):
        return redirect('/')
    users = load_json(USERS_FILE)
    if user_id in users:
        users[user_id]['is_admin'] = False
        save_json(USERS_FILE, users)
    return redirect('/admin')

@app.route('/admin/edit/<user_id>', methods=['GET', 'POST'])
def admin_edit(user_id):
    # Admin check
    if not session.get('user') or not session['user'].get('is_admin'):
        return redirect('/')

    users = load_json(USERS_FILE)
    if user_id not in users:
        return redirect('/admin')

    if request.method == 'POST':
        # ফর্ম থেকে ডাটা নাও
        name = request.form['name'].strip()
        email = request.form['email'].strip().lower()
        phone = request.form['phone'].strip()

        # ডাটা আপডেট করো
        users[user_id]['name'] = name or users[user_id].get('name')
        users[user_id]['email'] = email or users[user_id].get('email')
        users[user_id]['phone'] = phone or users[user_id].get('phone')

        save_json(USERS_FILE, users)

        # আপডেট হয়ে গেলে admin পেজে redirect করো
        return redirect('/admin')

    user = users[user_id]
    return render_template('admin_edit.html', user=user, user_id=user_id)
@app.route('/view_students')
def view_students():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    students = load_json(STUDENTS_FILE)
    user_students = [s for s in students if s.get('user_id') == user_id]

    batch = request.args.get('batch', '')
    if batch:
        user_students = [s for s in user_students if s.get('batch') == batch]

    # সব ব্যাচ ইউজারের ছাত্রদের থেকে নেয়া
    batches = sorted(set(s.get('batch') for s in user_students if s.get('batch')))

    return render_template('view_students.html', students=user_students, batches=batches, selected_batch=batch)


@app.route('/view_payments')
def view_payments():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    payments = load_json(PAYMENTS_FILE)
    user_payments = {roll: p for roll, p in payments.items() if p.get('user_id') == user_id}

    batch = request.args.get('batch', '')
    if batch:
        user_payments = {roll: p for roll, p in user_payments.items() if p.get('batch') == batch}

    # সব ব্যাচ ইউজারের payment এর থেকে নেয়া
    batches = sorted(set(p.get('batch') for p in user_payments.values() if p.get('batch')))

    return render_template('view_payments.html', payments=user_payments, batches=batches, selected_batch=batch)

@app.route('/delete_all_students', methods=['POST'])
def delete_all_students():
    if 'user' not in session:
        return redirect(url_for('login'))

    password = request.form.get('password')
    user = session['user']
    users = load_json(USERS_FILE)
    if not check_password_hash(users[user['email']]['password'], password):
        flash("❌ Incorrect password!", "danger")
        return redirect(url_for('dashboard'))

    all_students = load_json(STUDENTS_FILE)
    all_students = [s for s in all_students if s.get('user_id') != user['id']]
    save_json(STUDENTS_FILE, all_students)
    flash("✅ All students deleted successfully!", "success")
    return redirect(url_for('dashboard'))


@app.route('/delete_all_payments', methods=['POST'])
def delete_all_payments():
    if 'user' not in session:
        return redirect(url_for('login'))

    password = request.form.get('password')
    user = session['user']
    users = load_json(USERS_FILE)
    if not check_password_hash(users[user['email']]['password'], password):
        flash("❌ Incorrect password!", "danger")
        return redirect(url_for('payment'))

    all_payments = load_json(PAYMENTS_FILE)
    if user['id'] in all_payments:
        del all_payments[user['id']]
    save_json(PAYMENTS_FILE, all_payments)
    flash("✅ All payments deleted successfully!", "success")
    return redirect(url_for('payment'))


@app.route('/')
def home():
    if session.get('user'):
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


# ---------- RUN ----------
if __name__ == '__main__':
    app.run(debug=True)
